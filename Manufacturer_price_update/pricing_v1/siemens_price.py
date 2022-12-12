# importing modules
import openpyxl as xl
from tqdm import tqdm
from timeit import default_timer as timer


def Siemens_Price():

    # importing modules
    import openpyxl as xl
    from tqdm import tqdm
    from timeit import default_timer as timer

    # starting timer
    start = timer()

    print("Cadastramento de preços da Siemens selecionado")

    # opening the source excel file (official Siemens pricing list)
    # filename1 = input("Digite o caminho da lista de preços da Siemens: ")
    filename1 = "C:\\Users\\jamir\\OneDrive\\Área de Trabalho\\Scripts\\Manufacturer-price-update\\spreadsheets\\Siemens\\2019_01_lista_PV.xlsx"
    wb1 = xl.load_workbook(filename1, data_only=True, read_only=False)
    ws1 = wb1.active

    # Opening the destination excel file (agile software exported Siemens pricing list)
    # filename2 = input("Digite o caminho da lista a ser precificada: ")
    filename2 = "C:\\Users\\jamir\\OneDrive\\Área de Trabalho\\Scripts\\Manufacturer-price-update\\spreadsheets\\Siemens\\SIEMENS-Banco dados 18-10-22 sem desconto.xlsx"
    # Adapting to consider multiple worksheets inside one workbook
    n = 0
    wb2 = xl.load_workbook(filename2, data_only=True, read_only=False)
    sheets = wb2.sheetnames
    size = len(sheets)

    # opening the discount excel file
    # filename3 = input("Digite o caminho da lista de descontos Siemens: ")
    filename3 = "C:\\Users\\jamir\\OneDrive\\Área de Trabalho\\Scripts\\Manufacturer-price-update\\spreadsheets\\Siemens\\RelatorioDescontosCliente ENGEMAKRO 2016_2017 (8).xlsx"
    wb3 = xl.load_workbook(filename3, data_only=True, read_only=False)
    ws3 = wb3.active

    # calculate total number of rows in the source excel file
    mr1 = ws1.max_row

    # calculate total number of rows in the discount excel file
    mr3 = ws3.max_row

    CC1 = 2  # column where the component code is in the source excel file
    CC2 = 4  # column where the component code is in the destination excel file

    CP1 = 8  # column where the component price is in the source excel file
    CP2 = 7  # column where the component price is in the destination excel file

    CDCS = 5  # Column where the discount code is in the source excel file

    CDCD = 1  # Column where the discount code is in the discount excel file
    CDVD = 4  # Column where the discount multiplier is in the discount excel file

    mult = 1.6  # Base price multiplier (to make up for old pricing list)

    # i varies along the rows of the origin worksheet
    # j varies along the rows of the destination worksheet
    # y varies along the rows of the discount worksheet

    RepeatedSetIndex = set()

    for n in range(0, size):
        ws2 = wb2[sheets[n]]
        mr2 = ws2.max_row
        for j in tqdm(range(2, mr2 + 1)):
            for i in range(6, mr1 + 1):
                if i in RepeatedSetIndex:
                    continue
                codeSource = ws1.cell(row=i, column=CC1).value
                codeDestination = ws2.cell(row=j, column=CC2).value
                if codeSource == codeDestination:
                    RepeatedSetIndex.add(i)
                    discountCode1 = ws1.cell(row=i, column=CDCS).value
                    price = ws1.cell(row=i, column=CP1).value
                    priceCorrected = price * mult
                    if isinstance(price, str):
                        price = 0
                    ws2.cell(row=j, column=CP2).value = priceCorrected
                    for y in range(1, mr3 + 1):
                        discountCode2 = ws3.cell(row=y, column=CDCD).value
                        if discountCode1 == discountCode2:
                            discountValue = ws3.cell(row=y, column=CDVD).value
                            currentPrice = ws2.cell(row=j, column=CP2).value
                            ws2.cell(row=j, column=CP2).value = float(
                                currentPrice
                            ) * float(discountValue)

    # saving the destination excel file
    wb2.save(str(filename2))

    end = timer()
    print(
        f" Finalizado.\n Um total de {len(RepeatedSetIndex)} foram encontrados.\n Tempo total de processamento: {end - start} seconds\n"
    )
