# importing modules
import openpyxl as xl
from tqdm import tqdm
from timeit import default_timer as timer


def ABB_Price():

    # starting timer
    start = timer()

    print("Cadastramento de preços da ABB selecionado")

    # opening the source excel file (official ABB pricing list)
    # filename1 = input("Digite o caminho da lista de preços da ABB: ")
    filename1 = "C:\\Users\\jamir\\OneDrive\\Área de Trabalho\\Scripts\\Manufacturer-price-update\\spreadsheets\\ABB\\Oficial Lista de Preços Setembro 22.xlsx"
    wb1 = xl.load_workbook(filename1, data_only=True, read_only=False)
    ws1 = wb1.active

    # Opening the destination excel file (agile software exported pricing list)
    # filename2 = input("Digite o caminho da lista a ser precificada: ")
    filename2 = "C:\\Users\\jamir\\OneDrive\\Área de Trabalho\\Scripts\\Manufacturer-price-update\\spreadsheets\\ABB\\ABB-Banco Dados 12-10-2022.xlsx"
    # Adapting to consider multiple worksheets inside one workbook
    n = 0
    wb2 = xl.load_workbook(filename2, data_only=True, read_only=False)
    sheets = wb2.sheetnames
    size = len(sheets)

    # opening the ABB discount excel file
    # filename3 = input("Digite o caminho da lista de descontos ABB: ")
    filename3 = "C:\\Users\\jamir\\OneDrive\\Área de Trabalho\\Scripts\\Manufacturer-price-update\\spreadsheets\\ABB\\desconto ABB.xlsx"
    wb3 = xl.load_workbook(filename3, data_only=True, read_only=False)
    ws3 = wb3.active

    # calculate total number of rows in the source excel file
    mr1 = ws1.max_row

    # calculate total number of rows in the discount excel file
    mr3 = ws3.max_row

    CC1 = 6  # column where the component code is in the source excel file
    CC2 = 4  # column where the component code is in the destination excel file

    CP1 = 9  # column where the component price is in the source excel file
    CP2 = 7  # column where the component price is in the destination excel file

    CDCS = 5  # Column where the discount code is in the source excel file

    CDCD = 1  # Column where the discount code is in the discount excel file
    CDVD = 2  # Column where the discount multiplier is in the discount excel file

    mult = 1  # Base price multiplier (to make up for old pricing list)

    # i varies along the rows of the origin worksheet
    # j varies along the rows of the destination worksheet
    # y varies along the rows of the discount worksheet

    RepeatedSetIndex = set()
    count = 0

    for n in range(0, size):
        ws2 = wb2[sheets[n]]
        mr2 = ws2.max_row
        for j in tqdm(range(2, mr2 + 1)):
            for i in range(4, mr1 + 1):
                if i in RepeatedSetIndex:
                    count = count + 1
                    continue
                codeSource = ws1.cell(row=i, column=CC1).value
                codeDestination = ws2.cell(row=j, column=CC2).value
                if codeSource == codeDestination:
                    RepeatedSetIndex.add(i)
                    discountCode1 = ws1.cell(row=i, column=CDCS).value
                    price = ws1.cell(row=i, column=CP1).value
                    if isinstance(price, str):
                        price = 0
                    priceCorrected = price * mult
                    ws2.cell(row=j, column=CP2).value = priceCorrected
                    for y in range(0, mr3 + 1):
                        discountCode2 = ws3.cell(row=y, column=CDCD).value
                        if discountCode1 == discountCode2:
                            discountValue = ws3.cell(row=y, column=CDVD).value
                            currentPrice = ws2.cell(row=j, column=CP2).value
                            ws2.cell(row=j, column=CP2).value = float(
                                currentPrice
                            ) * float(discountValue)

    # saving the destination excel file
    wb2.save(str(filename2))

    end = timer()
    print(
        f" Finalizado.\n Um total de {len(RepeatedSetIndex)} foram encontrados.\n Tempo total de processamento: {end - start} segundos\n"
    )
    print(count)
